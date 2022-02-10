data = [
    ['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'],
    [1,10,8,5,14,26,12],
    [2,7,3,7,15,24,18],
    [3,9,5,8,8,12,4],
    [4,7,8,7,17,21,18],
    [5,7,8,7,16,25,15],
    [6,3,5,8,8,17,0],
    [7,4,9,10,16,27,18],
    [8,6,6,6,15,19,17],
    [9,10,10,9,19,30,19],
    [10,9,8,8,20,25,20],
]
# 비중
# 출석 10
# 퀴즈1 10
# 퀴즈2 10
# 중간고사 20
# 기말고사 30
# 프로젝트 20

# 1. 퀴즈2 점수를 10으로 수정
# 2. H열에 총점, I열에 성적 정보 추가
#     - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
# 3. 출석이 5 미만인 학생은 총점 상관없이 F

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

for d in data:
    ws.append(d)

for cell in ws['D'][1:]:
    cell.value = 10

ws['H1'] = '총점'
for idx, cell in enumerate(ws['H'][1:], start=2):
    cell.value = f'=SUM(B{idx}:G{idx})'

ws['I1'] = '성적 정보'
for idx, score in enumerate(data[1:], start=2):
    sum_val = sum(score[1:]) - score[3] + 10
    
    grade = None
    if sum_val >= 90:
        grade = 'A'
    elif sum_val >= 80:
        grade = 'B'
    elif sum_val >= 70:
        grade = 'C'
    else:
        grade = 'D'
    
    if score[1] < 5:
        grade = 'F'
    
    ws.cell(row=idx, column=9).value = grade

wb.save('scores.xlsx')
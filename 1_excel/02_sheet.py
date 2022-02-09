from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet
ws.title = 'MySheet'
ws.sheet_properties.tabColor = 'ff66ff' # 탭 색

ws1 = wb.create_sheet('YourSheet') # 주어진 이름으로 생성
ws2 = wb.create_sheet('NewSheet', 2) # 시트 인덱스

new_ws = wb['NewSheet']
print(wb.sheetnames) # 모든 시트명 확인

# 시트 복사
new_ws['A1'] = 'Test'
target = wb.copy_worksheet(new_ws)
target.title = 'Copied Sheet'
print(wb.sheetnames)

wb.save('sample.xlsx')
wb.close()
from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from random import *

wb = Workbook()
ws = wb.active

# 한줄씩 데이터 넣기
ws.append(['번호', '영어', '수학'])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws['B']
print(col_B)
for cell in col_B:
    print(cell.value, end=' ')

col_range = ws['B:C']
for cols in col_range:
    for cell in cols:
        print(cell.value, end=' ')
    print()

row_title = ws[1] # 첫번째 row만
for cell in row_title:
    print(cell.value)

row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value)

row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows: # 좌표정보
        # print(cell.value, end=' ')
        # print(cell.coordinate, end=' ')
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=' ')
    print()

# 전체
print(tuple(ws.rows))
print(tuple(ws.columns))

for row in tuple(ws.rows):
    print(row[1].value, end=' ')

for col in tuple(ws.columns):
    print(col[0].value, end=' ')

for row in ws.iter_rows():
    print(row[0].value)

for col in ws.iter_cols():
    print(col[0].value)

# 슬라이싱
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value)

wb.save('sample2.xlsx')
wb.close()
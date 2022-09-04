from openpyxl import load_workbook

wb = load_workbook('sample_merge.xlsx')
ws = wb.active

# 병합 해제
ws.unmerge_cells('B2:D2')
wb.save('sample_unmerge.xlsx')
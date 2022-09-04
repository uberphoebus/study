from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook('sample2.xlsx')
ws = wb.active

a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']

# 너비, 높이 조정
ws.column_dimensions['A'].width = 5
ws.row_dimensions[1].height = 30

# 스타일
a1.font = Font(color='FF0000', italic=True, bold=True)
b1.font = Font(color='CC33FF', name='Arial', strike=True)
c1.font = Font(color='0000FF', size=20, underline='single')

# 테두리
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 배경
for row in ws.rows:
    for cell in row:
        # 정렬
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if cell.column == 1: # A열 제외
            continue
        
        # 정수형 데이터, 90점 이상
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor='00FF00', fill_type='solid')
            cell.font = Font(color='FF0000')

# 틀 고정
ws.freeze_panes = 'B2'

wb.save('sample_style.xlsx')
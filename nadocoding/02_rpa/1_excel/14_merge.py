from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 셀 병합
ws.merge_cells('B2:D2')
ws['B2'].value = 'Merged Cell'

wb.save('sample_merge.xlsx')
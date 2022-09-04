from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'NadoSheet'

ws['A1'] = 1 # 값 지정
ws['A2'] = 2

ws['B1'] = 4
ws['B2'] = 6

print(ws['A1']) # 셀 정보
print(ws['A1'].value) # 셀 값
print(ws['A10'].value)

print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)

c = ws.cell(column=3, row=1, value=10) # 값 입력
print(c.value)

from random import *
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0, 100))

wb.save('sample1.xlsx')
wb.close()